# lrc_cross_screener_fast.py
# Binance SPOT + FUTURES(PERP) sadece LRC orta çizgi kesişmeleri (High LRC vs Low LRC)
# TF: 1D / 4H / 3H / 2H / 1H
# LRC_LEN = 298
# Excel: Coin / Timeframe / CrossType (renkli)
# Coin formatı:
#   - Spot:     XXXUSDT
#   - Perpetual:XXXUSDT.P
# Sadece USDT pariteleri taranır, alfabetik sıralanır.

import time
import math
from datetime import datetime, timezone

import numpy as np
import requests
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


# ===================== AYARLAR =====================
LRC_LEN = 298
LOOKBACK_DAYS = 30

TIMEFRAMES = {
    "1d": {"bin": "1d", "bars_per_day": 1},
    "4h": {"bin": "4h", "bars_per_day": 6},
    "3h": {"bin": "3h", "bars_per_day": 8},
    "2h": {"bin": "2h", "bars_per_day": 12},
    "1h": {"bin": "1h", "bars_per_day": 24},
}

SLEEP_BETWEEN_REQUESTS = 0.06
OUTPUT_FILE = "LRC_Kesisme_Sonuclari.xlsx"

# Stabil ve türev token filtreleri (istersen kapatabilirsin)
STABLE_BASES = {
    "USDC", "FDUSD", "TUSD", "USDP", "DAI", "BUSD", "EUR", "EURI",
    "UST", "USTC", "PAX", "PAXG",
}
BANNED_SUBSTRINGS = ("UPUSDT", "DOWNUSDT", "BULLUSDT", "BEARUSDT")

# ===================== HTTP =====================
session = requests.Session()
session.headers.update({"User-Agent": "Mozilla/5.0"})


def get_json(url, timeout=25, max_retries=5):
    backoff = 1.0
    for _ in range(max_retries):
        try:
            r = session.get(url, timeout=timeout)
            if r.status_code in (418, 429):
                time.sleep(backoff)
                backoff = min(backoff * 2.0, 15.0)
                continue
            r.raise_for_status()
            return r.json()
        except Exception:
            time.sleep(backoff)
            backoff = min(backoff * 2.0, 15.0)
    return None


# ===================== SYMBOL LISTS =====================
def get_spot_usdt_symbols():
    data = get_json("https://api.binance.com/api/v3/exchangeInfo", timeout=20)
    if not data:
        return []

    out = []
    for s in data.get("symbols", []):
        if s.get("status") != "TRADING":
            continue
        sym = s.get("symbol", "")
        if not sym.endswith("USDT"):
            continue
        if any(b in sym for b in BANNED_SUBSTRINGS):
            continue
        base = sym[:-4]
        if base in STABLE_BASES:
            continue
        out.append(sym)

    return sorted(set(out))


def get_futures_usdt_perp_symbols():
    data = get_json("https://fapi.binance.com/fapi/v1/exchangeInfo", timeout=20)
    if not data:
        return []

    out = []
    for s in data.get("symbols", []):
        if s.get("contractType") != "PERPETUAL":
            continue
        if s.get("status") != "TRADING":
            continue
        sym = s.get("symbol", "")
        if not sym.endswith("USDT"):
            continue
        if any(b in sym for b in BANNED_SUBSTRINGS):
            continue
        base = sym[:-4]
        if base in STABLE_BASES:
            continue
        out.append(sym)

    return sorted(set(out))


# ===================== KLINES =====================
def fetch_klines(symbol, interval, limit, futures=False):
    if futures:
        url = f"https://fapi.binance.com/fapi/v1/klines?symbol={symbol}&interval={interval}&limit={limit}"
    else:
        url = f"https://api.binance.com/api/v3/klines?symbol={symbol}&interval={interval}&limit={limit}"
    data = get_json(url, timeout=25)
    return data if isinstance(data, list) else []


# ===================== LRC (linreg end value) =====================
def rolling_linreg_end_values(arr, length):
    n = len(arr)
    out = np.full(n, np.nan)
    if n < length:
        return out

    x = np.arange(length, dtype=float)
    x_mean = x.mean()
    denom = ((x - x_mean) ** 2).sum()

    for i in range(length - 1, n):
        y = np.array(arr[i - length + 1: i + 1], dtype=float)
        y_mean = y.mean()
        cov = ((x - x_mean) * (y - y_mean)).sum()
        slope = cov / denom
        intercept = y_mean - slope * x_mean
        out[i] = slope * (length - 1) + intercept

    return out


def detect_crosses_lrc_mid(klines, lrc_len, lookback_bars):
    """
    TradingView karşılığı:
      lrcHighReg = ta.linreg(high, L, 0)
      lrcLowReg  = ta.linreg(low,  L, 0)
      crossover(lrcHighReg, lrcLowReg) / crossunder(...)
    """
    if not klines:
        return []

    highs = [float(k[2]) for k in klines]
    lows = [float(k[3]) for k in klines]
    times = [int(k[0]) for k in klines]  # open time ms

    high_lrc = rolling_linreg_end_values(highs, lrc_len)
    low_lrc = rolling_linreg_end_values(lows, lrc_len)

    n = len(highs)
    start = max(lrc_len, n - lookback_bars)
    crosses = []

    for i in range(start, n):
        if i - 1 < 0:
            continue
        if any(math.isnan(x) for x in (high_lrc[i], low_lrc[i], high_lrc[i - 1], low_lrc[i - 1])):
            continue

        prev_high, prev_low = high_lrc[i - 1], low_lrc[i - 1]
        curr_high, curr_low = high_lrc[i], low_lrc[i]

        if prev_high < prev_low and curr_high > curr_low:
            crosses.append({"ts": times[i], "type": "Crossover"})
        elif prev_high > prev_low and curr_high < curr_low:
            crosses.append({"ts": times[i], "type": "Crossunder"})

    return crosses


def iso_from_ms(ms):
    return datetime.fromtimestamp(ms / 1000, tz=timezone.utc).isoformat(timespec="seconds")


# ===================== SCAN =====================
def scan_markets():
    spot_symbols = get_spot_usdt_symbols()
    fut_symbols = get_futures_usdt_perp_symbols()

    print(f"SPOT USDT sembol: {len(spot_symbols)}")
    print(f"FUTURES PERP USDT sembol: {len(fut_symbols)}")

    results = []

    # Tek fonksiyonla iki marketi tara
    def scan_one_list(symbols, is_futures):
        mname = "FUT" if is_futures else "SPOT"
        total = len(symbols)
        for si, sym in enumerate(symbols, 1):
            base = sym[:-4]
            display = f"{base}USDT.P" if is_futures else f"{base}USDT"
            print(f"[{mname} {si}/{total}] {sym} taranıyor...")

            for tf_key, meta in TIMEFRAMES.items():
                interval = meta["bin"]
                lookback_bars = LOOKBACK_DAYS * meta["bars_per_day"]

                # Minimum limit: LRC_LEN + lookback + 5 (güven payı)
                limit_needed = LRC_LEN + lookback_bars + 5

                kl = fetch_klines(sym, interval, limit_needed, futures=is_futures)
                time.sleep(SLEEP_BETWEEN_REQUESTS)
                if not kl:
                    continue

                crosses = detect_crosses_lrc_mid(kl, LRC_LEN, lookback_bars)
                if not crosses:
                    continue

                # İstersen hepsini yaz, istersen sadece en son kesişme:
                # crosses_to_write = crosses[-1:]  # sadece son kesişme
                crosses_to_write = crosses        # son 30 gün içindeki tüm kesişmeler

                for c in crosses_to_write:
                    results.append({
                        "Coin": display,
                        "Timeframe": tf_key,
                        "CrossType": c["type"],
                        # İstemediğin için Excel'e koymuyoruz; debug için tutuyoruz:
                        "_CrossTime": iso_from_ms(c["ts"]),
                    })

    scan_one_list(spot_symbols, is_futures=False)
    scan_one_list(fut_symbols, is_futures=True)

    # alfabetik sıralama: Coin, sonra Timeframe, sonra CrossType
    tf_order = {"1d": 0, "4h": 1, "3h": 2, "2h": 3, "1h": 4}
    results.sort(key=lambda r: (r["Coin"], tf_order.get(r["Timeframe"], 99), r["CrossType"]))

    return results


# ===================== EXCEL =====================
def write_excel(rows, filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "LRC_Crosses"

    headers = ["Coin", "Timeframe", "CrossType"]
    ws.append(headers)

    # Renkler
    fill_cross_over = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # yeşil
    fill_cross_under = PatternFill(start_color="FFDDA0", end_color="FFDDA0", fill_type="solid") # turuncu

    for r in rows:
        ws.append([r["Coin"], r["Timeframe"], r["CrossType"]])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    col_crosstype = headers.index("CrossType") + 1

    for row_idx in range(2, ws.max_row + 1):
        v = ws.cell(row_idx, col_crosstype).value
        if v == "Crossover":
            ws.cell(row_idx, col_crosstype).fill = fill_cross_over
        elif v == "Crossunder":
            ws.cell(row_idx, col_crosstype).fill = fill_cross_under

    # genişlik
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 12

    wb.save(filename)
    print(f"Excel kaydedildi: {filename}")


# ===================== RUN =====================
def main():
    start = datetime.now()
    print("LRC hızlı tarama başladı:", start.isoformat(timespec="seconds"))

    rows = scan_markets()
    if rows:
        write_excel(rows, OUTPUT_FILE)
        print(f"Toplam {len(rows)} kesişme kaydı yazıldı.")
    else:
        print("Hiç kesişme bulunamadı.")

    end = datetime.now()
    print("Bitti:", end.isoformat(timespec="seconds"), "Süre(s):", (end - start).total_seconds())


if __name__ == "__main__":
    main()
